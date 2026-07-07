"""CRUD operations and exports for NBA staff records."""
import asyncio
import io
import logging

from fastapi import HTTPException

from app.config import settings
from app.db.postgres import get_pool
from app.services.storage_service import upload_staff_photo, upload_staff_qr, upload_staff_signature

logger = logging.getLogger(__name__)

_SELECT = (
    "id::text, full_name, department, photo_url, signature_url, qr_code_url, created_at"
)


# ── Internal helpers ──────────────────────────────────────────────────────────

async def _generate_and_store_qr(staff_id: str) -> str | None:
    """Generate QR PNG linking to the public staff profile, upload to GCS, update DB."""
    try:
        from app.services.qr_service import _generate_qr_png
        url = f"{settings.FRONTEND_ORIGIN}/staff/{staff_id}"
        png = await asyncio.to_thread(_generate_qr_png, url)
        qr_url = await upload_staff_qr(staff_id, png)
        pool = await get_pool()
        await pool.execute("UPDATE staff SET qr_code_url = $1 WHERE id = $2", qr_url, staff_id)
        return qr_url
    except Exception:
        logger.exception("Staff QR generation failed for id=%s", staff_id)
        return None


# ── Public API ────────────────────────────────────────────────────────────────

async def list_staff() -> list[dict]:
    pool = await get_pool()
    rows = await pool.fetch(
        f"SELECT {_SELECT} FROM staff ORDER BY full_name ASC"
    )
    return [dict(r) for r in rows]


async def get_staff_member(staff_id: str) -> dict:
    pool = await get_pool()
    row = await pool.fetchrow(f"SELECT {_SELECT} FROM staff WHERE id = $1", staff_id)
    if row is None:
        raise HTTPException(
            status_code=404,
            detail={"code": "STAFF_NOT_FOUND", "message": "Staff member not found.", "details": {}},
        )
    return dict(row)


async def get_public_profile(staff_id: str) -> dict:
    """Minimal public profile — no auth required."""
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT id::text, full_name, department, photo_url, qr_code_url FROM staff WHERE id = $1",
        staff_id,
    )
    if row is None:
        raise HTTPException(
            status_code=404,
            detail={"code": "STAFF_NOT_FOUND", "message": "Staff profile not found.", "details": {}},
        )
    return dict(row)


async def create_staff(full_name: str, department: str) -> dict:
    pool = await get_pool()
    row = await pool.fetchrow(
        f"INSERT INTO staff (full_name, department) VALUES ($1, $2) RETURNING {_SELECT}",
        full_name.strip(),
        department.strip(),
    )
    staff = dict(row)
    qr_url = await _generate_and_store_qr(staff["id"])
    if qr_url:
        staff["qr_code_url"] = qr_url
    return staff


async def regenerate_qr(staff_id: str) -> dict:
    await get_staff_member(staff_id)  # 404 if missing
    await _generate_and_store_qr(staff_id)
    return await get_staff_member(staff_id)


async def set_staff_photo(staff_id: str, data: bytes, content_type: str) -> dict:
    await get_staff_member(staff_id)
    url = await upload_staff_photo(staff_id, data, content_type)
    pool = await get_pool()
    row = await pool.fetchrow(
        f"UPDATE staff SET photo_url = $1 WHERE id = $2 RETURNING {_SELECT}", url, staff_id
    )
    return dict(row)


async def set_staff_signature(staff_id: str, data: bytes, content_type: str) -> dict:
    await get_staff_member(staff_id)
    url = await upload_staff_signature(staff_id, data, content_type)
    pool = await get_pool()
    row = await pool.fetchrow(
        f"UPDATE staff SET signature_url = $1 WHERE id = $2 RETURNING {_SELECT}", url, staff_id
    )
    return dict(row)


async def delete_staff(staff_id: str) -> None:
    await get_staff_member(staff_id)
    pool = await get_pool()
    await pool.execute("DELETE FROM staff WHERE id = $1", staff_id)


# ── Excel export ──────────────────────────────────────────────────────────────

async def export_staff_xlsx() -> bytes:
    """Build an .xlsx workbook with all staff: photo, name, department, signature, QR code."""
    import httpx

    rows = await list_staff()

    async def _fetch(client: httpx.AsyncClient, url: str | None) -> io.BytesIO | None:
        if not url:
            return None
        try:
            r = await client.get(url, timeout=8.0)
            if r.status_code == 200:
                buf = io.BytesIO(r.content)
                buf.seek(0)
                return buf
        except Exception:
            pass
        return None

    async with httpx.AsyncClient() as client:
        photo_bufs, sig_bufs = await asyncio.gather(
            asyncio.gather(*[_fetch(client, r.get("photo_url")) for r in rows]),
            asyncio.gather(*[_fetch(client, r.get("signature_url")) for r in rows]),
        )

    return await asyncio.to_thread(_build_staff_xlsx, rows, list(photo_bufs), list(sig_bufs))


def _build_staff_xlsx(
    rows: list[dict],
    photo_bufs: list[io.BytesIO | None],
    sig_bufs: list[io.BytesIO | None],
) -> bytes:
    from concurrent.futures import ThreadPoolExecutor
    from io import BytesIO as _BytesIO

    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from PIL import Image as PILImage

    from app.services.qr_service import _generate_qr_png

    QR_DATA_PX = 200
    QR_DISP_PX = 80
    PHOTO_DATA_W, PHOTO_DATA_H = 810, 1080
    PHOTO_DISP_W, PHOTO_DISP_H = 90, 120
    SIG_DISP_W, SIG_DISP_H = 150, 50
    ROW_H = 120

    def _render_qr(staff_id: str) -> "_BytesIO | None":
        try:
            raw = _generate_qr_png(f"{settings.FRONTEND_ORIGIN}/staff/{staff_id}")
            pil = PILImage.open(_BytesIO(raw)).resize((QR_DATA_PX, QR_DATA_PX), PILImage.LANCZOS)
            buf = _BytesIO()
            pil.save(buf, "PNG")
            buf.seek(0)
            return buf
        except Exception:
            logger.exception("Staff QR render failed for id=%s", staff_id)
            return None

    def _render_photo(src: "_BytesIO | None") -> "_BytesIO | None":
        if src is None:
            return None
        try:
            src.seek(0)
            pil = PILImage.open(src).convert("RGB")
            orig_w, orig_h = pil.size
            scale = max(PHOTO_DATA_W / orig_w, PHOTO_DATA_H / orig_h)
            new_w, new_h = round(orig_w * scale), round(orig_h * scale)
            pil = pil.resize((new_w, new_h), PILImage.LANCZOS)
            left = (pil.width - PHOTO_DATA_W) // 2
            top = (pil.height - PHOTO_DATA_H) // 2
            pil = pil.crop((left, top, left + PHOTO_DATA_W, top + PHOTO_DATA_H))
            buf = _BytesIO()
            pil.save(buf, "JPEG", quality=85, optimize=True)
            buf.seek(0)
            return buf
        except Exception:
            logger.exception("Staff photo render failed")
            return None

    def _render_sig(src: "_BytesIO | None") -> "_BytesIO | None":
        if src is None:
            return None
        try:
            src.seek(0)
            pil = PILImage.open(src).convert("RGBA")
            # Scale to fixed height, preserve aspect ratio
            ratio = SIG_DISP_H / pil.height
            new_w = max(1, round(pil.width * ratio * 4))  # 4× for print quality
            new_h = SIG_DISP_H * 4
            pil = pil.resize((new_w, new_h), PILImage.LANCZOS)
            bg = PILImage.new("RGB", (new_w, new_h), (255, 255, 255))
            bg.paste(pil, mask=pil.split()[3] if pil.mode == "RGBA" else None)
            buf = _BytesIO()
            bg.save(buf, "PNG")
            buf.seek(0)
            return buf
        except Exception:
            logger.exception("Staff signature render failed")
            return None

    with ThreadPoolExecutor(max_workers=8) as pool:
        qr_bufs = list(pool.map(_render_qr, [r["id"] for r in rows]))
        photo_rendered = list(pool.map(_render_photo, photo_bufs))
        sig_rendered = list(pool.map(_render_sig, sig_bufs))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff"

    headers = ["Photo", "Full Name", "Department", "Signature", "QR Code"]
    ws.append(headers)

    green_fill = PatternFill("solid", fgColor="1A5C2A")
    white_bold = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    col_widths = [14, 30, 32, 22, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for idx, (row, qr_buf, photo_buf, sig_buf) in enumerate(
        zip(rows, qr_bufs, photo_rendered, sig_rendered), start=2
    ):
        ws.append(["", row["full_name"], row["department"], "", ""])
        ws.row_dimensions[idx].height = ROW_H

        for col in range(1, 6):
            ws.cell(row=idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)

        if photo_buf:
            photo_buf.seek(0)
            xl_img = XLImage(photo_buf)
            xl_img.width = PHOTO_DISP_W
            xl_img.height = PHOTO_DISP_H
            ws.add_image(xl_img, f"A{idx}")

        if sig_buf:
            sig_buf.seek(0)
            xl_sig = XLImage(sig_buf)
            xl_sig.width = SIG_DISP_W
            xl_sig.height = SIG_DISP_H
            ws.add_image(xl_sig, f"D{idx}")

        if qr_buf:
            qr_buf.seek(0)
            xl_qr = XLImage(qr_buf)
            xl_qr.width = QR_DISP_PX
            xl_qr.height = QR_DISP_PX
            ws.add_image(xl_qr, f"E{idx}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
