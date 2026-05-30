"""Admin service — member directory, stats, status management, and audit logging.

All DB operations use asyncpg directly.
"""
import asyncio
import csv
import io
import json
import logging
import re
import secrets
import string

import asyncpg
import httpx
from fastapi import HTTPException, status

from app.config import settings
from app.db.postgres import get_pool
from app.models.profile import ProfileCreate
from app.services import photo_service, qr_service, storage_service

logger = logging.getLogger(__name__)

_SORT_FIELDS = {"created_at", "full_name", "branch"}
_ENROLLMENT_RE = re.compile(r"^[A-Za-z0-9/\-]+$")


# ── Async public API ───────────────────────────────────────────────────────────

async def get_stats() -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        counts = await conn.fetchrow(
            """SELECT
               COUNT(*)                                          AS total,
               COUNT(*) FILTER (WHERE status = 'active')        AS active_count,
               COUNT(*) FILTER (WHERE status = 'pending')       AS pending_count,
               COUNT(*) FILTER (WHERE status = 'suspended')     AS suspended_count,
               COUNT(*) FILTER (WHERE payment_status = 'paid')  AS paid_count,
               COUNT(*) FILTER (WHERE payment_status = 'unpaid') AS unpaid_count
               FROM public.member_profiles"""
        )
        latest_row = await conn.fetchrow(
            "SELECT full_name, created_at FROM public.member_profiles ORDER BY created_at DESC LIMIT 1"
        )
        branch_rows = await conn.fetch(
            "SELECT branch, COUNT(*) AS cnt FROM public.member_profiles GROUP BY branch ORDER BY cnt DESC"
        )

    latest_member = None
    if latest_row:
        latest_member = {
            "full_name": latest_row["full_name"],
            "created_at": str(latest_row["created_at"]),
        }

    return {
        "total_members": counts["total"],
        "active_members": counts["active_count"],
        "pending_members": counts["pending_count"],
        "suspended_members": counts["suspended_count"],
        "paid_members": counts["paid_count"],
        "unpaid_members": counts["unpaid_count"],
        "latest_member": latest_member,
        "members_by_branch": [
            {"branch": r["branch"], "count": r["cnt"]} for r in branch_rows
        ],
    }


async def list_members(
    q: str | None = None,
    status_filter: str | None = None,
    branch: str | None = None,
    year_of_call: int | None = None,
    payment_status: str | None = None,
    page: int = 1,
    page_size: int = 50,
    sort_by: str = "created_at",
    sort_dir: str = "desc",
) -> tuple[list[dict], int]:
    page_size = min(page_size, 200)
    if sort_by not in _SORT_FIELDS:
        sort_by = "created_at"
    offset = (page - 1) * page_size
    order = "DESC" if sort_dir == "desc" else "ASC"

    conditions: list[str] = []
    params: list = []
    idx = 1

    if q:
        conditions.append(
            f"(full_name ILIKE ${idx} OR branch ILIKE ${idx} OR enrollment_no ILIKE ${idx})"
        )
        params.append(f"%{q}%")
        idx += 1
    if status_filter:
        conditions.append(f"status = ${idx}")
        params.append(status_filter)
        idx += 1
    if branch:
        conditions.append(f"branch = ${idx}")
        params.append(branch)
        idx += 1
    if year_of_call is not None:
        conditions.append(f"year_of_call = ${idx}")
        params.append(year_of_call)
        idx += 1
    if payment_status:
        conditions.append(f"payment_status = ${idx}")
        params.append(payment_status)
        idx += 1

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    params.extend([page_size, offset])

    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT * FROM public.member_profiles {where} "
            f"ORDER BY {sort_by} {order} LIMIT ${idx} OFFSET ${idx + 1}",
            *params,
        )
        total_row = await conn.fetchrow(
            f"SELECT COUNT(*) AS cnt FROM public.member_profiles {where}",
            *params[:-2],  # exclude LIMIT/OFFSET params
        )

    return [dict(r) for r in rows], total_row["cnt"]


async def get_member_detail(member_id: str) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT * FROM public.member_profiles WHERE id = $1",
            member_id,
        )
        if not profile:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")
        payments = await conn.fetch(
            "SELECT * FROM public.payment_transactions WHERE member_id = $1 ORDER BY created_at DESC",
            member_id,
        )

    return {**dict(profile), "payment_history": [dict(p) for p in payments]}


async def update_status(
    admin_id: str,
    member_id: str,
    new_status: str,
    reason: str | None,
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT * FROM public.member_profiles WHERE id = $1",
            member_id,
        )
        if not profile:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")

        old_status = profile["status"]

        async with conn.transaction():
            updated = await conn.fetchrow(
                "UPDATE public.member_profiles SET status = $1 WHERE id = $2 RETURNING *",
                new_status,
                member_id,
            )
            await conn.execute(
                """INSERT INTO public.admin_audit_log
                   (admin_id, action, target_id, old_value, new_value)
                   VALUES ($1, $2, $3, $4::jsonb, $5::jsonb)""",
                admin_id,
                "status_change",
                member_id,
                json.dumps({"status": old_status}),
                json.dumps({"status": new_status, "reason": reason}),
            )

    return dict(updated)


async def regenerate_qr(admin_id: str, member_id: str) -> str | None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT id FROM public.member_profiles WHERE id = $1",
            member_id,
        )
    if not profile:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")

    url = await qr_service.generate_and_store(member_id)

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """INSERT INTO public.admin_audit_log
               (admin_id, action, target_id, old_value, new_value)
               VALUES ($1, $2, $3, NULL, $4::jsonb)""",
            admin_id,
            "qr_regenerated",
            member_id,
            json.dumps({"qr_code_url": url}),
        )

    return url


async def update_enrollment_no(
    admin_id: str,
    member_id: str,
    new_enrollment_no: str,
) -> dict:
    """Update a member's enrollment number (SCN). Validates format, checks uniqueness,
    and logs the change to admin_audit_log.

    Raises:
        HTTPException 400 — invalid enrollment number format.
        HTTPException 404 — member not found.
        HTTPException 409 — enrollment number already in use.
    """
    new_enrollment_no = new_enrollment_no.strip().upper()
    if not _ENROLLMENT_RE.match(new_enrollment_no):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "code": "VALIDATION_ERROR",
                "message": "Enrollment number may only contain letters, digits, '/' and '-'.",
                "details": {},
            },
        )

    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT id, enrollment_no FROM public.member_profiles WHERE id = $1",
            member_id,
        )
        if not profile:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")

        old_enrollment_no = profile["enrollment_no"]

        try:
            async with conn.transaction():
                updated = await conn.fetchrow(
                    "UPDATE public.member_profiles SET enrollment_no = $1 WHERE id = $2 RETURNING *",
                    new_enrollment_no,
                    member_id,
                )
                await conn.execute(
                    """INSERT INTO public.admin_audit_log
                       (admin_id, action, target_id, old_value, new_value)
                       VALUES ($1, $2, $3, $4::jsonb, $5::jsonb)""",
                    admin_id,
                    "enrollment_no_change",
                    member_id,
                    json.dumps({"enrollment_no": old_enrollment_no}),
                    json.dumps({"enrollment_no": new_enrollment_no}),
                )
        except asyncpg.UniqueViolationError:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="DUPLICATE_ENROLLMENT",
            )

    return dict(updated)


_UPDATABLE_FIELDS = frozenset({
    "full_name", "year_of_call", "branch", "phone_number",
    "email_address", "office_address", "status", "payment_status",
})


async def update_profile(admin_id: str, member_id: str, updates: dict) -> dict:
    """Update any subset of a member's profile fields. Logs to admin_audit_log."""
    fields = {k: v for k, v in updates.items() if v is not None and k in _UPDATABLE_FIELDS}
    if not fields:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="NO_FIELDS_TO_UPDATE")

    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT * FROM public.member_profiles WHERE id = $1", member_id
        )
        if not profile:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")

        old_values = {k: profile[k] for k in fields if k in profile.keys()}

        set_clauses = []
        params: list = []
        idx = 1
        for col, val in fields.items():
            set_clauses.append(f"{col} = ${idx}")
            params.append(val)
            idx += 1
        params.append(member_id)

        async with conn.transaction():
            updated = await conn.fetchrow(
                f"UPDATE public.member_profiles SET {', '.join(set_clauses)} "
                f"WHERE id = ${idx} RETURNING *",
                *params,
            )
            await conn.execute(
                """INSERT INTO public.admin_audit_log
                   (admin_id, action, target_id, old_value, new_value)
                   VALUES ($1, $2, $3, $4::jsonb, $5::jsonb)""",
                admin_id,
                "profile_updated",
                member_id,
                json.dumps({k: str(v) for k, v in old_values.items()}),
                json.dumps({k: str(v) for k, v in fields.items()}),
            )

    return dict(updated)


async def delete_member(admin_id: str, member_id: str) -> None:
    """Delete a member profile and log the action. Auth account is left intact."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT id, full_name FROM public.member_profiles WHERE id = $1", member_id
        )
        if not profile:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")

        async with conn.transaction():
            await conn.execute(
                "DELETE FROM public.member_profiles WHERE id = $1", member_id
            )
            await conn.execute(
                """INSERT INTO public.admin_audit_log
                   (admin_id, action, target_id, old_value, new_value)
                   VALUES ($1, $2, $3, $4::jsonb, NULL)""",
                admin_id,
                "profile_deleted",
                member_id,
                json.dumps({"full_name": profile["full_name"]}),
            )


async def get_vcard(member_id: str) -> str:
    pool = await get_pool()
    async with pool.acquire() as conn:
        profile = await conn.fetchrow(
            "SELECT * FROM public.member_profiles WHERE id = $1",
            member_id,
        )
    if not profile:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="PROFILE_NOT_FOUND")
    return _format_vcard(dict(profile))


async def export_csv(
    status_filter: str | None = None,
    branch: str | None = None,
    payment_status: str | None = None,
) -> str:
    conditions: list[str] = []
    params: list = []
    idx = 1

    if status_filter:
        conditions.append(f"status = ${idx}")
        params.append(status_filter)
        idx += 1
    if branch:
        conditions.append(f"branch = ${idx}")
        params.append(branch)
        idx += 1
    if payment_status:
        conditions.append(f"payment_status = ${idx}")
        params.append(payment_status)
        idx += 1

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT * FROM public.member_profiles {where} ORDER BY created_at DESC",
            *params,
        )

    return _format_csv([dict(r) for r in rows])


async def get_audit_log(page: int = 1, page_size: int = 50) -> tuple[list[dict], int]:
    page_size = min(page_size, 200)
    offset = (page - 1) * page_size

    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM public.admin_audit_log ORDER BY created_at DESC LIMIT $1 OFFSET $2",
            page_size,
            offset,
        )
        total_row = await conn.fetchrow("SELECT COUNT(*) AS cnt FROM public.admin_audit_log")

    return [dict(r) for r in rows], total_row["cnt"]


async def create_member(
    member_id: str,
    data: ProfileCreate,
    photo_bytes: bytes | None,
    mime: str | None,
) -> dict:
    """Create a profile for an existing auth user with payment waived (status=active)."""
    photo_url: str | None = None

    if photo_bytes is not None and mime is not None:
        stage2 = await photo_service.validate_photo_stage2(photo_bytes, mime)
        if not stage2.passed:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "PHOTO_REJECTED",
                    "message": "Your photo did not pass compliance checks.",
                    "details": {"failures": stage2.failures, "score": stage2.score},
                },
            )
        photo_url = await storage_service.upload_photo(member_id, photo_bytes, mime)

    # Collision-free member_uid generation
    uid: str | None = None
    for _ in range(5):
        candidate = _generate_member_uid()
        pool = await get_pool()
        async with pool.acquire() as conn:
            existing = await conn.fetchrow(
                "SELECT id FROM public.member_profiles WHERE member_uid = $1",
                candidate,
            )
        if not existing:
            uid = candidate
            break
    if uid is None:
        raise HTTPException(status_code=500, detail="INTERNAL_ERROR")

    profile_url = f"{settings.FRONTEND_ORIGIN}/profile/{uid}"

    pool = await get_pool()
    async with pool.acquire() as conn:
        try:
            row = await conn.fetchrow(
                """INSERT INTO public.member_profiles
                   (id, full_name, enrollment_no, year_of_call, branch,
                    phone_number, email_address, office_address,
                    photo_url, member_uid, profile_url, status, payment_status)
                   VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13)
                   RETURNING *""",
                member_id,
                data.full_name,
                data.enrollment_no,
                data.year_of_call,
                data.branch,
                data.phone_number,
                str(data.email_address),
                data.office_address,
                photo_url,
                uid,
                profile_url,
                "active",
                "paid",
            )
        except asyncpg.UniqueViolationError as exc:
            constraint = exc.constraint_name or ""
            if "enrollment_no" in constraint:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="DUPLICATE_ENROLLMENT")
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A profile already exists for this account.",
            )
        except Exception as exc:
            logger.error("Admin profile insert failed: %s", exc)
            raise HTTPException(status_code=500, detail="INTERNAL_ERROR")

    asyncio.create_task(qr_service.generate_and_store(member_id))
    return dict(row)


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _generate_member_uid() -> str:
    chars = string.ascii_uppercase + string.digits
    part1 = "".join(secrets.choice(chars) for _ in range(6))
    part2 = "".join(secrets.choice(chars) for _ in range(8))
    return f"NBA-{part1}-{part2}"


def _format_vcard(profile: dict) -> str:
    return (
        "BEGIN:VCARD\r\n"
        "VERSION:3.0\r\n"
        f"FN:{profile['full_name']}\r\n"
        "ORG:Nigerian Bar Association\r\n"
        f"TEL;TYPE=CELL:{profile['phone_number']}\r\n"
        f"EMAIL:{profile['email_address']}\r\n"
        f"ADR:;;{profile['office_address']};;;;\r\n"
        f"NOTE:NBA Member - {profile['member_uid']} | Branch: {profile['branch']} | "
        f"Year of Call: {profile['year_of_call']}\r\n"
        "END:VCARD\r\n"
    )


def _format_csv(members: list[dict]) -> str:
    buf = io.StringIO()
    fields = [
        "member_uid", "full_name", "branch", "year_of_call", "enrollment_no",
        "email_address", "phone_number", "status", "payment_status", "created_at",
    ]
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(members)
    return buf.getvalue()


def _format_sheets_csv(members: list[dict]) -> str:
    """CSV with =IMAGE() formulas for QR/photo columns — import directly into Google Sheets."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Full Name", "Year of Call", "Branch", "QR Code", "Passport Photo"])
    for m in members:
        qr = f'=IMAGE("{m["qr_code_url"]}")' if m.get("qr_code_url") else ""
        photo = f'=IMAGE("{m["photo_url"]}")' if m.get("photo_url") else ""
        writer.writerow([
            m.get("full_name", ""),
            m.get("year_of_call", ""),
            m.get("branch", ""),
            qr,
            photo,
        ])
    return buf.getvalue()


async def _fetch_photo_buf(client: httpx.AsyncClient, url: str | None) -> io.BytesIO | None:
    if not url:
        return None
    try:
        r = await client.get(url, timeout=8.0)
        if r.status_code == 200:
            buf = io.BytesIO(r.content)
            buf.seek(0)
            return buf
    except Exception:
        pass
    return None


def _build_xlsx(rows: list[dict], photo_bufs: list[io.BytesIO | None]) -> bytes:
    """Build an .xlsx workbook with member data, QR codes, and passport photos embedded as images.

    QR PNGs are generated in parallel (up to 8 threads); photo bytes are
    pre-fetched asynchronously and passed in.
    """
    from concurrent.futures import ThreadPoolExecutor
    from io import BytesIO as _BytesIO

    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from PIL import Image as PILImage

    from app.services.qr_service import _generate_qr_png

    # Image data resolution (high-res for extraction) vs display size in the sheet
    QR_DATA_PX = 200     # embedded PNG resolution
    QR_DISP_PX = 80      # displayed cell size (pts ≈ px at 96dpi)
    PHOTO_DATA_W, PHOTO_DATA_H = 810, 1080   # embedded PNG — 1080p portrait
    PHOTO_DISP_W, PHOTO_DISP_H = 90, 120    # displayed cell size
    ROW_H = 120  # row height in points, matches PHOTO_DISP_H

    def _render_qr(m: dict) -> "_BytesIO | None":
        uid = m.get("member_uid")
        if not uid:
            return None
        try:
            raw = _generate_qr_png(f"{settings.FRONTEND_ORIGIN}/profile/{uid}")
            pil = PILImage.open(_BytesIO(raw)).resize((QR_DATA_PX, QR_DATA_PX), PILImage.LANCZOS)
            buf = _BytesIO()
            pil.save(buf, "PNG")
            buf.seek(0)
            return buf
        except Exception:
            logger.exception("QR render failed for member_uid=%s", uid)
            return None

    def _render_photo(src: "_BytesIO | None") -> "_BytesIO | None":
        if src is None:
            return None
        try:
            src.seek(0)
            pil = PILImage.open(src).convert("RGB")
            orig_w, orig_h = pil.size
            target_w, target_h = PHOTO_DATA_W, PHOTO_DATA_H
            # Scale so the shorter axis meets the target; crop to exact size.
            # Works for both upscaling (small photos) and downscaling (large photos).
            scale = max(target_w / orig_w, target_h / orig_h)
            new_w, new_h = round(orig_w * scale), round(orig_h * scale)
            pil = pil.resize((new_w, new_h), PILImage.LANCZOS)
            left = (pil.width - target_w) // 2
            top = (pil.height - target_h) // 2
            pil = pil.crop((left, top, left + target_w, top + target_h))
            buf = _BytesIO()
            pil.save(buf, "PNG")
            buf.seek(0)
            return buf
        except Exception:
            logger.exception("Photo render failed")
            return None

    with ThreadPoolExecutor(max_workers=8) as pool:
        qr_bufs = list(pool.map(_render_qr, rows))
        photo_resized = list(pool.map(_render_photo, photo_bufs))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = [
        "Full Name", "Branch", "Year of Call", "Enrollment No.",
        "Email", "Phone", "Status", "Payment", "QR Code", "Photo",
    ]
    ws.append(headers)

    green_fill = PatternFill("solid", fgColor="1A5C2A")
    white_bold = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    col_widths = [28, 14, 12, 16, 28, 16, 10, 10, 12, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    for row_num, (m, qr_buf, photo_buf) in enumerate(zip(rows, qr_bufs, photo_resized), start=2):
        ws.append([
            m.get("full_name", ""),
            m.get("branch", ""),
            m.get("year_of_call", ""),
            m.get("enrollment_no", ""),
            m.get("email_address", ""),
            m.get("phone_number", ""),
            m.get("status", ""),
            m.get("payment_status", ""),
            "", "",  # placeholders for QR (I) and Photo (J)
        ])
        ws.row_dimensions[row_num].height = ROW_H

        if qr_buf:
            xl_img = XLImage(qr_buf)
            xl_img.width = QR_DISP_PX
            xl_img.height = QR_DISP_PX
            ws.add_image(xl_img, f"I{row_num}")

        if photo_buf:
            xl_img = XLImage(photo_buf)
            xl_img.width = PHOTO_DISP_W
            xl_img.height = PHOTO_DISP_H
            ws.add_image(xl_img, f"J{row_num}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def export_xlsx(
    status_filter: str | None = None,
    branch: str | None = None,
    payment_status: str | None = None,
    limit: int | None = None,
    offset: int = 0,
) -> bytes:
    conditions: list[str] = []
    params: list = []
    idx = 1

    if status_filter:
        conditions.append(f"status = ${idx}")
        params.append(status_filter)
        idx += 1
    if branch:
        conditions.append(f"branch = ${idx}")
        params.append(branch)
        idx += 1
    if payment_status:
        conditions.append(f"payment_status = ${idx}")
        params.append(payment_status)
        idx += 1

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    limit_clause = ""
    if limit is not None and limit > 0:
        limit_clause = f" LIMIT ${idx} OFFSET ${idx + 1}"
        params.append(min(limit, 5000))
        params.append(max(offset, 0))

    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT * FROM public.member_profiles {where} ORDER BY created_at DESC{limit_clause}",
            *params,
        )

    row_dicts = [dict(r) for r in rows]
    logger.info("export_xlsx: fetching photos for %d members", len(row_dicts))

    async with httpx.AsyncClient() as client:
        photo_bufs = list(await asyncio.gather(*[
            _fetch_photo_buf(client, m.get("photo_url")) for m in row_dicts
        ]))

    logger.info("export_xlsx: building workbook")
    try:
        return await asyncio.to_thread(_build_xlsx, row_dicts, photo_bufs)
    except Exception:
        logger.exception("export_xlsx: _build_xlsx failed")
        raise


async def export_sheets_csv(
    status_filter: str | None = None,
    branch: str | None = None,
    limit: int | None = None,
) -> str:
    conditions: list[str] = []
    params: list = []
    idx = 1

    if status_filter:
        conditions.append(f"status = ${idx}")
        params.append(status_filter)
        idx += 1
    if branch:
        conditions.append(f"branch = ${idx}")
        params.append(branch)
        idx += 1

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    limit_clause = ""
    if limit is not None and limit > 0:
        limit_clause = f" LIMIT ${idx}"
        params.append(min(limit, 5000))

    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"SELECT full_name, year_of_call, branch, qr_code_url, photo_url "
            f"FROM public.member_profiles {where} ORDER BY created_at DESC{limit_clause}",
            *params,
        )

    return _format_sheets_csv([dict(r) for r in rows])
